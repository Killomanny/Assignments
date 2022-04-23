import pandas as pd
import openpyxl
# from openpyxl.styles import Alignment
from collections import Counter

df = pd.read_excel('raw.xlsx', 'Sheet1', header=None)
a = df.to_numpy()

df = pd.DataFrame(a)

newlst = []

df.applymap(lambda x: newlst.append(x))

newlst.sort()

cou = Counter(newlst)
marks = list(cou.keys())
freq = list(cou.values())
lt = len(marks)
marks_freq = []
i = 0
while i != lt:
    marks_freq.append(marks[i]*freq[i])
    i = i+1


total = len(newlst)

df1 = pd.DataFrame()
df1['marks'] = marks
df1['No of students (freq)'] = freq
df1['freq times mark (fi times marks )'] = marks_freq
df1.to_excel('out.xlsx', index=False)


wb = openpyxl.load_workbook("out.xlsx")
ws = wb['Sheet1']
ws['A15'] = 'Total'
ws['B15'] = total
ws['C15'] = sum(marks_freq)
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 30
ws.row_dimensions[1].height = 30

print("sum total =",sum(marks_freq))
print("total no of students =",total)
print("Mean is =",sum(marks_freq)/total)


wb.save("out.xlsx")
wb.close()
